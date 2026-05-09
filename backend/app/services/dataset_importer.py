"""
Dataset label importers — convert various formats to canonical training JSON.

See SPEC.md §9.2.
M3.5 milestone.

Speaker convention:
  - Internal: 1-indexed int (matches user-facing display)
  - Training JSON output: 0-indexed int (matches upstream toy_dataset)
  → conversion happens at import time (decrement by 1)
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

from app.constants import DATASET_LABEL_FORMATS
from app.errors import AppError, ErrorCode

logger = logging.getLogger(__name__)


# ============================================================
# Time parsing
# ============================================================


_TIME_RE_HMS = re.compile(r"^(\d+):(\d{1,2}):(\d{1,2})(?:\.(\d+))?$")
_TIME_RE_MS = re.compile(r"^(\d+):(\d{1,2})(?:\.(\d+))?$")


def parse_time(value: Any) -> float:
    """
    Parse time as float seconds or hh:mm:ss[.ms] or mm:ss[.ms].
    Raises AppError(IMPORT_INVALID_TIME) on bad input.
    """
    if isinstance(value, int | float):
        return float(value)
    s = str(value).strip()
    if not s:
        raise AppError(ErrorCode.IMPORT_INVALID_TIME, "Empty time")
    # Try plain float first
    try:
        return float(s)
    except ValueError:
        pass
    # hh:mm:ss
    if m := _TIME_RE_HMS.match(s):
        h, mi, sec, ms = m.groups()
        return int(h) * 3600 + int(mi) * 60 + int(sec) + (float(f"0.{ms}") if ms else 0)
    # mm:ss
    if m := _TIME_RE_MS.match(s):
        mi, sec, ms = m.groups()
        return int(mi) * 60 + int(sec) + (float(f"0.{ms}") if ms else 0)
    raise AppError(ErrorCode.IMPORT_INVALID_TIME, f"Cannot parse time: {s!r}")


def parse_speaker(value: Any) -> int:
    """
    Parse speaker. Accepts:
      - int (0 or 1+): used as-is, but must be >=0
      - "Speaker 1" / "Sp1" / "S1" / "speaker 0": extract digit, return 0-indexed
      - " 1 " (str digit): used as-is

    NOTE: returned value is the value the USER provided semantically. Caller
    should normalize to 0-indexed for training JSON output.
    """
    if isinstance(value, int | float):
        return int(value)
    s = str(value).strip()
    if s.isdigit():
        return int(s)
    m = re.search(r"\d+", s)
    if m:
        n = int(m.group())
        # If user wrote "Speaker 1", they mean speaker index 1 (1-based).
        # We assume convention: any string form is 1-indexed → convert to 0-indexed.
        return max(0, n - 1)
    return 0


# ============================================================
# Format-specific parsers
# ============================================================


def parse_xlsx(label_path: Path) -> list[dict]:
    """
    Parse Excel file. 第一個 sheet，header 必須含
    start_time / end_time / speaker / text（順序不限）。
    回傳 list[{start, end, speaker, text}]，speaker 已 0-indexed。
    """
    from openpyxl import load_workbook

    wb = load_workbook(label_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise AppError(ErrorCode.IMPORT_PARSE_FAILED, "No active sheet")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise AppError(ErrorCode.IMPORT_PARSE_FAILED, "Empty xlsx")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    required = ("start_time", "end_time", "speaker", "text")
    missing = [k for k in required if k not in header]
    if missing:
        raise AppError(
            ErrorCode.IMPORT_PARSE_FAILED, f"Missing column(s): {missing}"
        )
    idx = {k: header.index(k) for k in required}

    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        try:
            start = parse_time(row[idx["start_time"]])
            end = parse_time(row[idx["end_time"]])
        except AppError:
            raise  # IMPORT_INVALID_TIME 直接往上拋
        speaker = parse_speaker(row[idx["speaker"]])
        text_val = row[idx["text"]]
        text = "" if text_val is None else str(text_val).strip()
        if not text:
            continue
        out.append({"start": start, "end": end, "speaker": speaker, "text": text})
    return out


def parse_srt(label_path: Path) -> list[dict]:
    """
    Parse SubRip file. Speaker prefix optional ('Speaker N: ...' / 'Speaker N：...').
    無 prefix 時 speaker = 0；有 prefix 時 N - 1 → 0-indexed。
    """
    import pysrt

    subs = pysrt.open(str(label_path), encoding="utf-8")
    out: list[dict] = []
    for sub in subs:
        start = sub.start.ordinal / 1000.0
        end = sub.end.ordinal / 1000.0
        text = sub.text.strip() if sub.text else ""
        m = re.match(r"^Speaker\s*(\d+)\s*[:：]\s*(.*)$", text, re.DOTALL)
        if m:
            speaker = max(0, int(m.group(1)) - 1)
            text = m.group(2).strip()
        else:
            speaker = 0
        out.append({"start": start, "end": end, "speaker": speaker, "text": text})
    return out


def parse_json(label_path: Path) -> list[dict]:
    """Parse JSON file (assumed already in training format). Pass-through with validation."""
    data = json.loads(label_path.read_text(encoding="utf-8"))
    segs = data.get("segments")
    if not isinstance(segs, list):
        raise AppError(ErrorCode.IMPORT_PARSE_FAILED, "JSON missing 'segments' list")
    return segs


def parse_txt(label_path: Path, audio_duration: float = 0.0) -> list[dict]:
    """
    Parse plain text. 每行：[hh:mm:ss.ms] Speaker N: text
    end_time 推算：下一行 start；最後一行用 audio_duration。
    """
    line_re = re.compile(
        r"^\[(\d+:\d{1,2}:\d{1,2}(?:\.\d+)?)\]\s+Speaker\s*(\d+)\s*[:：]\s*(.+)$"
    )
    raw: list[tuple[float, int, str]] = []
    for line in label_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        m = line_re.match(line)
        if not m:
            logger.warning("parse_txt: skip unmatched line: %r", line[:80])
            continue
        ts, sp, txt = m.groups()
        start = parse_time(ts)
        speaker = max(0, int(sp) - 1)
        raw.append((start, speaker, txt.strip()))

    if not raw:
        raise AppError(ErrorCode.IMPORT_PARSE_FAILED, "No valid lines parsed")

    out: list[dict] = []
    for i, (start, speaker, text) in enumerate(raw):
        end = raw[i + 1][0] if i + 1 < len(raw) else audio_duration
        out.append({"start": start, "end": end, "speaker": speaker, "text": text})
    return out


# ============================================================
# Top-level entry
# ============================================================


def import_label(
    label_path: Path,
    audio_filename: str,
    audio_duration: float,
    format: str,
    project_hotwords: list[str],
) -> dict:
    """Top-level: parse → validate → return canonical training JSON dict."""
    if format not in DATASET_LABEL_FORMATS:
        raise AppError(
            ErrorCode.UNSUPPORTED_FORMAT,
            f"Format must be one of {list(DATASET_LABEL_FORMATS)}",
        )
    if format == "txt":
        segments = parse_txt(label_path, audio_duration=audio_duration)
    elif format == "json":
        segments = parse_json(label_path)
    elif format == "xlsx":
        segments = parse_xlsx(label_path)
    else:  # srt（前面 set 已 guard，此處為 exhaustive 收口）
        segments = parse_srt(label_path)
    validate_segments(segments, audio_duration)
    return {
        "audio_duration": audio_duration,
        "audio_path": audio_filename,
        "segments": segments,
        "customized_context": project_hotwords,
    }


def validate_segments(segments: list[dict], audio_duration: float) -> None:
    """
    Sanity checks on parsed segments.
    Raises AppError(IMPORT_PARSE_FAILED) on hard errors.
    """
    if not segments:
        raise AppError(ErrorCode.IMPORT_PARSE_FAILED, "No segments parsed")
    prev_end = -1.0
    for i, seg in enumerate(segments):
        for k in ("speaker", "text", "start", "end"):
            if k not in seg:
                raise AppError(
                    ErrorCode.IMPORT_PARSE_FAILED,
                    f"Segment {i}: missing key {k!r}",
                )
        if seg["start"] >= seg["end"]:
            raise AppError(
                ErrorCode.IMPORT_PARSE_FAILED,
                f"Segment {i}: start >= end ({seg['start']}, {seg['end']})",
            )
        if seg["start"] < prev_end - 0.01:  # allow tiny float jitter
            raise AppError(
                ErrorCode.IMPORT_PARSE_FAILED,
                f"Segment {i}: overlaps previous (start {seg['start']} < prev_end {prev_end})",
            )
        if seg["end"] > audio_duration + 0.5:
            raise AppError(
                ErrorCode.IMPORT_PARSE_FAILED,
                f"Segment {i}: end {seg['end']} exceeds audio duration {audio_duration}",
            )
        prev_end = seg["end"]
