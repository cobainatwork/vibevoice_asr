import io
import json
import pytest
from openpyxl import load_workbook
from app.services.dataset_exporter import export_item


class _FakeItem:
    """模擬 DatasetItem ORM row（exporter 純函式不 sticky DB）。"""
    def __init__(self, id: int, label: dict):
        self.id = id
        self.label = label


_LABEL = {
    "audio_duration": 7.20,
    "audio_path": "audio.wav",
    "segments": [
        {"speaker": 0, "text": "早安", "start": 0.0, "end": 3.45},
        {"speaker": 1, "text": "你好", "start": 3.45, "end": 7.20},
    ],
    "customized_context": ["糖尿病"],
}


def test_export_json_returns_bytes_and_correct_mime():
    item = _FakeItem(12, _LABEL)
    content, ct, name = export_item(item, "json")
    assert ct == "application/json"
    assert name == "dataset-12.json"
    parsed = json.loads(content.decode("utf-8"))
    # JSON 不轉 speaker（保持 0-indexed）
    assert parsed["segments"][0]["speaker"] == 0


def test_export_srt_speaker_one_indexed():
    item = _FakeItem(12, _LABEL)
    content, ct, name = export_item(item, "srt")
    assert ct == "application/x-subrip"
    assert name == "dataset-12.srt"
    text = content.decode("utf-8")
    # SRT speaker 顯示為 1-indexed
    assert "Speaker 1: 早安" in text
    assert "Speaker 2: 你好" in text
    assert "00:00:00,000 --> 00:00:03,450" in text


def test_export_xlsx_round_trip():
    """生 xlsx → openpyxl 讀回 → speaker 是 1-indexed、時間 hh:mm:ss.cc。"""
    item = _FakeItem(12, _LABEL)
    content, ct, name = export_item(item, "xlsx")
    assert ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert name == "dataset-12.xlsx"
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row 0 = header
    assert rows[0] == ("start_time", "end_time", "speaker", "text")
    assert rows[1][2] == 1  # 0-indexed → 1-indexed
    assert rows[2][2] == 2
    assert rows[1][3] == "早安"


def test_export_xlsx_time_format_hms_cc():
    item = _FakeItem(12, _LABEL)
    content, _, _ = export_item(item, "xlsx")
    wb = load_workbook(io.BytesIO(content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    # 0.00 sec → "00:00:00.00"
    assert rows[1][0] == "00:00:00.00"
    # 3.45 sec → "00:00:03.45"
    assert rows[1][1] == "00:00:03.45"


def test_export_unknown_format_raises():
    item = _FakeItem(12, _LABEL)
    with pytest.raises(ValueError):
        export_item(item, "csv")  # type: ignore[arg-type]


def test_export_json_indent_and_unicode():
    item = _FakeItem(12, _LABEL)
    content, _, _ = export_item(item, "json")
    text = content.decode("utf-8")
    # 不轉 ASCII（中文不變 \uXXXX）
    assert "早安" in text
    # 有縮排
    assert "  " in text
